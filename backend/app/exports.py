# -*- coding: utf-8 -*-
"""تصدير البيانات إلى CSV و Excel بترميز يدعم العربية."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def to_csv(headers: list[str], rows: list[list]) -> bytes:
    """CSV بترميز UTF-8 مع BOM ليُفتح بالعربية في Excel مباشرة."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    return ("﻿" + buf.getvalue()).encode("utf-8")


def to_xlsx(title: str, headers: list[str], rows: list[list], text_columns: set[int] | None = None) -> bytes:
    """ملف Excel منسّق مع رأس ملوّن.

    text_columns: فهارس أعمدة (0-based) تُفرض كنص صراحة (تنسيق '@') — تمنع تحويل
    Excel أرقام هوية طويلة (الرقم المدني) تلقائيًا لصيغة علمية مثل 1E+11 (QA-P1-RPT-01).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Sheet"
    ws.sheet_view.rightToLeft = True  # اتجاه عربي
    header_fill = PatternFill("solid", fgColor="0E5A54")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in rows:
        ws.append(r)
    if text_columns:
        for col_idx in text_columns:
            col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
            for row_idx in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row_idx}"]
                cell.number_format = "@"
                if cell.value is not None:
                    cell.value = str(cell.value)
    # عرض الأعمدة تلقائيًا (تقديري)
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width + 4, 40)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
